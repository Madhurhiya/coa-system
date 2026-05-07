# coa/management/commands/import_master_data.py
# Uses ONLY openpyxl + built-in csv — NO pandas required

import os
import csv
import openpyxl
from django.core.management.base import BaseCommand
from coa.models import ItemMaster, Customer

CAT_COL_MAP = {
    'Essential Oil':  'Essential Oil – Part Used',
    'Fixed Oil':      'Oil Soluble Extract – Part Used',
    'Oil Soluble':    'Oil Soluble Extract – Part Used',
    'Water Soluble':  'Water Soluble Extract – Part Used',
    'Dry Extract':    'Dry Extract – Part Used',
    'Soft Extract':   'Dry Extract – Part Used',
}


def clean(val):
    if val is None:
        return ''
    s = str(val).strip()
    return '' if s in ['–', '-', 'None', 'nan', ''] else s


class Command(BaseCommand):
    help = 'Import Item Master (Excel) and Customers (CSV)'

    def add_arguments(self, parser):
        parser.add_argument('--items', default='Item_Master_Enriched.xlsx')
        parser.add_argument('--customers', default='Contacts__2_.csv')

    def handle(self, *args, **options):
        base       = os.path.dirname(os.path.abspath('manage.py'))
        items_path = os.path.join(base, options['items'])
        cust_path  = os.path.join(base, options['customers'])

        if not os.path.exists(items_path):
            self.stdout.write(self.style.ERROR(f'Cannot find: {items_path}'))
            return
        if not os.path.exists(cust_path):
            self.stdout.write(self.style.ERROR(f'Cannot find: {cust_path}'))
            return

        # Import Items from Excel
        self.stdout.write('Reading Excel file...')
        wb = openpyxl.load_workbook(items_path, read_only=True, data_only=True)
        ws = wb.active

        headers = [clean(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col = {h: i for i, h in enumerate(headers) if h}

        created = updated = skipped = 0
        for row_values in ws.iter_rows(min_row=2, values_only=True):
            cat  = clean(row_values[col.get('Item Category', 0)]) or 'Other'
            name = clean(row_values[col.get('Item Name', 1)])
            if not name:
                skipped += 1
                continue
            botanical    = clean(row_values[col.get('Botanical Name(s)', 2)])
            general_part = clean(row_values[col.get('Plant Part(s) Used', 3)])
            specific_col  = CAT_COL_MAP.get(cat)
            specific_part = ''
            if specific_col and specific_col in col:
                specific_part = clean(row_values[col[specific_col]])
            plant_part = specific_part or general_part
            _, was_created = ItemMaster.objects.update_or_create(
                item_name=name,
                defaults={'item_category': cat, 'botanical_name': botanical, 'plant_part': plant_part}
            )
            if was_created: created += 1
            else: updated += 1
        wb.close()
        self.stdout.write(self.style.SUCCESS(f'Items: {created} created, {updated} updated, {skipped} skipped'))

        # Import Customers from CSV
        self.stdout.write('Reading CSV file...')
        c_created = c_skip = 0
        with open(cust_path, encoding='utf-8', newline='') as f:
            for row in csv.DictReader(f):
                name = row.get('Display Name', '').strip()
                if not name:
                    continue
                _, was_created = Customer.objects.get_or_create(name=name)
                if was_created: c_created += 1
                else: c_skip += 1

        self.stdout.write(self.style.SUCCESS(f'Customers: {c_created} imported, {c_skip} already existed'))
        self.stdout.write(self.style.SUCCESS('Import complete!'))